from openpyxl import *
from selenium import webdriver
import time
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By 
import openpyxl
from selenium import webdriver
import pandas as pd
import time
import re
import random
from openpyxl import Workbook
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from bs4 import BeautifulSoup
import warnings 

def start_code():
    search_key = "한국 저출산"    # 검색할 단어
    excelfile = load_workbook('test5.xlsx')    # 엑셀 파일 로드
    count = 0    # 영상 갯수 확인용

    # n번째 크롤링일 때----
    excelfile = load_workbook('test5.xlsx')    # text.xlsx 파일 로드
    sheet = excelfile[search_key]    # 시트 로드

    # D열의 값을 읽어올 리스트
    video_names = []

    # D열의 값 읽기 (D1은 제외)
    for row in sheet.iter_rows(min_row=2, values_only=True, max_col=4):
        video_name = row[3]  # D열은 0-based 인덱스로 3
        if video_name and video_name not in video_names:  # 값이 비어있지 않은 경우에만 리스트에 추가
            video_names.append(video_name)

    # 결과 출력 (옵션)
    print(video_names)

    #driver = webdriver.Chrome('C:/Users/ledu2/Python/Crawling/chromedriver.exe')
    driver = webdriver.Chrome()    # 현재 폴더에 chromedriver.exe가 있어서 경로 설정X
    #driver.get('https://www.youtube.com/results?search_query='+search_key)    # 그냥 유튜브 검색
    driver.get('https://www.youtube.com/@newskbs/search?query='+search_key)    # 뉴스 페이지에서 검색
    time.sleep(3)

    #html = driver.find_element_by_tag_name('html')    # 이거 못씀. 문법 바뀜
    html = driver.find_element(By.TAG_NAME, "html")    # 이거로 바뀜
    prev_contents_len = 0 

    try:
        while True: 
            time.sleep(3) 
            #contents = driver.find_elements(By.XPATH, '//ytd-page-manager[@id="page-manager"]/ytd-search/div[@id="container"]/ytd-two-column-search-results-renderer/div[@id="primary"]/ytd-section-list-renderer/div[@id="contents"]/ytd-item-section-renderer/div[@id="contents"]/ytd-video-renderer') 
            #contents = driver.find_elements(By.XPATH, '//*[@id="contents"]/ytd-video-renderer')    # 뉴스 페이지에서 검색
            contents = driver.find_elements(By.XPATH, '//*[@id="dismissible"]/ytd-thumbnail')    # 썸네일
            current_contents_len = len(contents)
            print(current_contents_len)
            if current_contents_len >= 100:    # 영상이 100개 이상 로드되면 중지
                break 
            html.send_keys(Keys.END)

        driver.execute_script('window.scrollTo(0,0)')

        #for _ in range(3):
        for _ in range(current_contents_len):
            time.sleep(3)
            warnings.filterwarnings('ignore')

            video_date = {}

            date_pattern = re.compile('[가-힣\s:]')
            print(len(contents))

            before_location2 = driver.execute_script("return window.pageYOffset")
            driver.execute_script("window.scrollTo(0,{})".format(before_location2 + 100))
            for idx, c in enumerate(contents):
                print(c)
                count+=1
                c.click()
                del contents[0]    # 봤던 영상은 리스트에서 제거
                time.sleep(3)

                try:
                    driver.find_element(
                        By.XPATH, '//*[@id="dismiss-button"]/yt-button-renderer/yt-button-shape/button/yt-touch-feedback-shape/div/div[2]').click()
                except:
                    pass
                time.sleep(3)

                #try:
                driver.execute_script("window.scrollTo(0,200)")
                time.sleep(2)
                try:
                    more_button = driver.find_element(By.XPATH, '//*[@id="snippet"]')    # 더보기 버튼 클릭
                    more_button.click()
                except:
                    driver.back()

                time.sleep(2)

                date = date_pattern.sub("", driver.find_element(By.XPATH, '//*[@id="info"]/span[3]').text)
                split_date = date.split(".")
                year, month = split_date[0], split_date[1]
                time.sleep(1)

                if date not in video_date.keys():
                    video_date[date] = 0
                else:
                    video_date[date] += 1

                if video_date[date] > 5:
                    continue

                title = driver.find_element(By.XPATH, '//*[@id="title"]/h1/yt-formatted-string').text
                if title in video_names:    # 이미 크롤링 했던 영상이라면 패스하도록
                    driver.back()
                    continue
                print(title + "/" + str(date) + "/" + str(video_date[date]))

                html = driver.find_element(By.TAG_NAME, "html")    # 이거로 바뀜
                html.send_keys(Keys.END)
                time.sleep(3)

                before_location = driver.execute_script("return window.pageYOffset")
                #while True:
                for _ in range(150):
                    #현재 위치 + 500으로 스크롤 이동
                    driver.execute_script("window.scrollTo(0,{})".format(before_location + 500))

                    #전체 스크롤이 늘어날 때까지 대기
                    time.sleep(3)

                    #이동 후 스크롤 위치
                    after_location = driver.execute_script("return window.pageYOffset")

                    #이동후 위치와 이동 후 위치가 같으면(더 이상 스크롤이 늘어나지 않으면) 종료
                    if before_location == after_location:
                        break

                    #같지 않으면 다음 조건 실행
                    else:
                        #이동여부 판단 기준이 되는 이전 위치 값 수정
                        before_location = driver.execute_script("return window.pageYOffset")

                try:
                    driver.find_element(By.CSS_SELECTOR, "#dismiss-button > a").click()
                except:
                    pass

                # 대댓글
                buttons = driver.find_elements(By.CSS_SELECTOR, "#more-replies > a")
                #more-replies > yt-button-shape > button

                time.sleep(3)

                for button in buttons:
                    button.send_keys(Keys.ENTER)
                    time.sleep(2.5)
                    button.click()
                #여기까지

                html_source = driver.page_source
                soup = BeautifulSoup(html_source, 'html.parser')

                id_list = soup.select("div#header-author > h3 > #author-text > span")
                comment_list = soup.select("yt-formatted-string#content-text")

                comment_final = []

                driver.back()
                for i in range(len(comment_list)):
                    temp_comment = comment_list[i].text
                    temp_comment = temp_comment.replace('\n', '')
                    temp_comment = temp_comment.replace('\t', '')
                    temp_comment = temp_comment.replace('    ', '')
                    print(temp_comment)
                    comment_final.append(temp_comment) # 댓글 내용

                for idx, comment in enumerate(comment_final): 
                    sheet.append([count, year, month, title, comment]) 
                    excelfile.save('./test5.xlsx') 
                    print(idx)
    except Exception as ex:
        print(ex)
        driver.quit()

for _ in range(100):
    start_code()